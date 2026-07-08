from datetime import datetime, date, timedelta
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Account, Category, Transaction, EntryType

EXCEL_EPOCH = datetime(1899, 12, 30)
REQUIRED_HEADERS = {"日期", "類別", "項目", "金額"}


def _to_clean_str(value) -> str | None:
    """儲存格內容可能是 str/int/float(使用者誤填數字),統一轉字串再 strip。"""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _excel_serial_to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # 可能是使用者手動打成 YYYYMMDD 整數格式(例如 20260708),而非真正的 Excel 日期
    if isinstance(value, (int, float)) and float(value) == int(value):
        ival = int(value)
        if 19000101 <= ival <= 99991231:
            try:
                return datetime.strptime(str(ival), "%Y%m%d").date()
            except ValueError:
                pass  # 不是合法日期(例如 20261332),繼續往下當 Excel 序列值處理
    # 一般情況:Excel 序列值(距離 1899-12-30 的天數)
    return (EXCEL_EPOCH + timedelta(days=float(value))).date()


def _find_header_row(ws: Worksheet) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """掃描前 10 列,找出同時包含「日期/類別/項目/金額」四個標題文字的列,
    回傳 (標題列列號, {標題文字: 欄位號}),沒找到回傳 (None, None)。"""
    scan_rows = min(10, ws.max_row)
    for row in ws.iter_rows(min_row=1, max_row=scan_rows):
        found: dict[str, int] = {}
        for cell in row:
            label = _to_clean_str(cell.value)
            if label in REQUIRED_HEADERS and label not in found:
                found[label] = cell.column
        if len(found) == len(REQUIRED_HEADERS):
            return row[0].row, found
    return None, None


def parse_month_sheets(file_bytes: bytes) -> list[dict]:
    """掃描所有工作表,自動偵測含「日期/類別/項目/金額」標題列的表(不限工作表名稱/欄位排列),
    找不到完整標題列的工作表(如總表、類別下拉選單來源)自動跳過。"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx, col_map = _find_header_row(ws)
        if header_row_idx is None:
            continue

        date_col = col_map["日期"]
        category_col = col_map["類別"]
        item_col = col_map["項目"]
        amount_col = col_map["金額"]

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=date_col)
            category_cell = ws.cell(row=row_idx, column=category_col)
            item_cell = ws.cell(row=row_idx, column=item_col)
            amount_cell = ws.cell(row=row_idx, column=amount_col)

            if category_cell.value is None and amount_cell.value is None:
                continue

            rows.append({
                "sheet": sheet_name,
                "row": row_idx,
                "date_raw": date_cell.value,
                "category_top": _to_clean_str(category_cell.value),
                "item": _to_clean_str(item_cell.value),
                "amount_raw": amount_cell.value,
            })
    return rows


def _validate_row(raw: dict) -> tuple[dict | None, str | None]:
    if not raw["category_top"]:
        return None, "缺少類別"
    if raw["amount_raw"] is None:
        return None, "缺少金額"
    try:
        amount = float(raw["amount_raw"])
    except (TypeError, ValueError):
        return None, f"金額格式錯誤: {raw['amount_raw']!r}"
    if amount <= 0:
        return None, f"金額必須為正數: {amount}"
    try:
        tx_date = _excel_serial_to_date(raw["date_raw"])
    except (TypeError, ValueError, OverflowError):
        return None, f"日期格式錯誤: {raw['date_raw']!r}"

    entry_type = EntryType.income if raw["category_top"] == "收入" else EntryType.expense
    return {
        "sheet": raw["sheet"], "row": raw["row"], "date": tx_date,
        "category_top": raw["category_top"], "item": raw["item"],
        "amount": amount, "type": entry_type,
    }, None


def _get_or_create_top_category(db: Session, household_id: str, name: str, type_: EntryType, dry_run: bool) -> Category | None:
    cat = db.execute(
        select(Category).where(
            Category.household_id == household_id,
            Category.parent_id.is_(None),
            Category.name == name,
            Category.type == type_,
        )
    ).scalar_one_or_none()
    if cat or dry_run:
        return cat
    cat = Category(household_id=household_id, name=name, parent_id=None, type=type_)
    db.add(cat)
    db.flush()
    return cat


def _find_category_anywhere(db: Session, household_id: str, name: str, type_: EntryType) -> Category | None:
    """不限層級比對同名分類,相容既有多層樹,避免重複建立。"""
    return db.execute(
        select(Category).where(
            Category.household_id == household_id,
            Category.name == name,
            Category.type == type_,
        )
    ).scalars().first()


def _existing_tx_keys(db: Session, household_id: str, account_id: str, tx_dates: set[date]) -> set[tuple]:
    if not tx_dates:
        return set()
    rows = db.execute(
        select(Transaction.date, Transaction.amount, Transaction.category_id, Transaction.type)
        .where(
            Transaction.household_id == household_id,
            Transaction.account_id == account_id,
            Transaction.date.in_(tx_dates),
        )
    ).all()
    return {(r.date, float(r.amount), r.category_id, r.type) for r in rows}


def process_import(
    db: Session, household_id: str, user_id: str, account_id: str,
    file_bytes: bytes, dry_run: bool, force_rows: set[str] | None = None,
) -> dict:
    force_rows = force_rows or set()
    raw_rows = parse_month_sheets(file_bytes)
    errors: list[str] = []
    new_categories: set[str] = set()
    preview_rows: list[dict] = []
    imported = 0
    skipped_duplicates = 0
    created_categories = 0
    balance_delta = 0.0  # 累計這批匯入對帳戶餘額的淨影響(收入+/支出-)

    top_cache: dict[tuple[str, EntryType], Category | None] = {}
    item_cache: dict[tuple[str, EntryType], Category | None] = {}

    parsed_list: list[dict] = []
    for raw in raw_rows:
        parsed, err = _validate_row(raw)
        if err:
            errors.append(f"[{raw['sheet']} row {raw['row']}] {err}")
            continue
        parsed_list.append(parsed)

    existing_keys = _existing_tx_keys(db, household_id, account_id, {p["date"] for p in parsed_list})
    seen_in_batch: set[tuple] = set()

    for parsed in parsed_list:
        top_key = (parsed["category_top"], parsed["type"])
        top_cat = top_cache.get(top_key)
        if top_cat is None and top_key not in top_cache:
            top_cat = _get_or_create_top_category(db, household_id, parsed["category_top"], parsed["type"], dry_run)
            top_cache[top_key] = top_cat
            if top_cat is None:
                new_categories.add(parsed["category_top"])
                created_categories += 1 if not dry_run else 0

        category_id_for_tx = top_cat.id if top_cat else None
        will_create = top_cat is None

        # 項目一律建成子分類(含卡費 > 富邦/兆豐...)
        if parsed["item"]:
            item_key = (parsed["item"], parsed["type"])
            item_cat = item_cache.get(item_key)
            if item_cat is None and item_key not in item_cache:
                item_cat = _find_category_anywhere(db, household_id, parsed["item"], parsed["type"])
                if item_cat is None and top_cat is not None and not dry_run:
                    item_cat = Category(
                        household_id=household_id, name=parsed["item"],
                        parent_id=top_cat.id, type=parsed["type"],
                    )
                    db.add(item_cat)
                    db.flush()
                    created_categories += 1
                if item_cat is None:
                    new_categories.add(f"{parsed['category_top']} > {parsed['item']}")
                    will_create = True
                item_cache[item_key] = item_cat
            if item_cat is not None:
                category_id_for_tx = item_cat.id

        row_key = f"{parsed['sheet']}:{parsed['row']}"
        dedupe_key = (parsed["date"], parsed["amount"], category_id_for_tx, parsed["type"])
        is_duplicate = dedupe_key in existing_keys or dedupe_key in seen_in_batch
        forced = row_key in force_rows

        preview_rows.append({
            "sheet": parsed["sheet"], "row": parsed["row"], "date": parsed["date"],
            "category_top": parsed["category_top"], "item": parsed["item"],
            "amount": parsed["amount"], "type": parsed["type"].value,
            "will_create_category": will_create,
            "is_duplicate": is_duplicate, "error": None,
        })

        if is_duplicate and not forced:
            skipped_duplicates += 1
            continue

        if not forced:
            seen_in_batch.add(dedupe_key)
        if not dry_run and category_id_for_tx:
            db.add(Transaction(
                household_id=household_id, user_id=user_id, account_id=account_id,
                category_id=category_id_for_tx, amount=parsed["amount"],
                type=parsed["type"], date=parsed["date"], note=None,
            ))
            imported += 1
            if parsed["type"] == EntryType.income:
                balance_delta += parsed["amount"]
            else:
                balance_delta -= parsed["amount"]

    if not dry_run:
        if balance_delta != 0:
            account = db.get(Account, account_id)
            account.balance = float(account.balance) + balance_delta
        db.commit()

    return {
        "total_rows": len(raw_rows),
        "valid_rows": len(preview_rows),
        "new_categories": sorted(new_categories),
        "errors": errors,
        "rows": preview_rows,
        "imported": imported,
        "skipped_errors": len(errors),
        "skipped_duplicates": skipped_duplicates,
        "created_categories": created_categories,
    }


def _root_category_name(db: Session, category: Category, cache: dict[str, Category]) -> str:
    node = category
    while node.parent_id is not None:
        parent = cache.get(node.parent_id)
        if parent is None:
            parent = db.get(Category, node.parent_id)
            cache[node.parent_id] = parent
        node = parent
    return node.name


def build_export_workbook(db: Session, household_id: str, year: int) -> BytesIO:
    txs = db.execute(
        select(Transaction)
        .where(Transaction.household_id == household_id)
        .where(Transaction.date >= date(year, 1, 1), Transaction.date <= date(year, 12, 31))
        .order_by(Transaction.date)
    ).scalars().all()

    cat_cache: dict[str, Category] = {}
    by_month: dict[int, list[Transaction]] = {m: [] for m in range(1, 13)}
    for tx in txs:
        by_month[tx.date.month].append(tx)

    wb = Workbook()
    wb.remove(wb.active)

    for month in range(1, 13):
        ws = wb.create_sheet(f"{month}月")
        ws.append(["日期", "類別", "項目", "金額"])
        for tx in by_month[month]:
            category = cat_cache.get(tx.category_id)
            if category is None:
                category = db.get(Category, tx.category_id)
                cat_cache[tx.category_id] = category
            root_name = _root_category_name(db, category, cat_cache)
            item = category.name if category.parent_id is not None else (tx.note or "")
            ws.append([tx.date, root_name, item, float(tx.amount)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
